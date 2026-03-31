import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _style_header_row(ws, num_cols: int):
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws, min_width=15, max_width=60):
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length, min_width), max_width)


def _wrap_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_controls_excel(
    document_name: str,
    controls: list[dict],
    gap_findings: list[dict],
    recommendations: list[dict],
    output_dir: str = ".",
    repo_controls: list[dict] = None,
) -> str:
    """
    Create an Excel workbook with three tabs:
    - Draft Controls (risk events on the left, associated controls to the right)
    - Gap Analysis
    - Recommendations

    Returns the path to the saved file.
    """
    wb = Workbook()

    # Build a lookup: control_id -> control dict (extracted controls)
    controls_by_id = {c.get("control_id"): c for c in controls}
    # Build a lookup: protecht_id -> repo control dict (repository controls)
    repo_by_id = {c.get("protecht_id"): c for c in (repo_controls or []) if c.get("protecht_id")}

    # -------------------------------------------------------------------------
    # Tab 1: Draft Controls — risk events leftmost, controls to the right
    # -------------------------------------------------------------------------
    ws_controls = wb.active
    ws_controls.title = "Draft Controls"

    draft_headers = [
        "Control ID", "Control Type", "Control Title", "Control Description",
        "Control Owner", "Control Performer", "Frequency", "Page", "Section",
        "Source Excerpt", "Expected Evidence", "Repository Match", "Protecht ID", "Control Status",
    ]
    ws_controls.append(draft_headers)

    for finding in gap_findings:
        controls_str = finding.get("controls_addressing", "None")

        associated_ids = [
            c.strip() for c in controls_str.split(",")
            if c.strip() and c.strip().lower() != "none"
        ]

        # Only include net-new and partial-match document controls.
        # Exact repository matches are already approved — they appear on the
        # Gap Analysis tab and do not need drafting.
        doc_ctrl_ids = [
            cid for cid in associated_ids
            if cid in controls_by_id
            and controls_by_id[cid].get("repository_match_type") != "exact"
        ]

        for ctrl_id in doc_ctrl_ids:
            ctrl = controls_by_id[ctrl_id]
            match_type = ctrl.get("repository_match_type", "")
            protecht_id = ctrl.get("repository_protecht_id", "")
            match_label = {"exact": "Existing", "partial": "Partial"}.get(match_type, "New")
            ws_controls.append([
                ctrl.get("control_id", ctrl_id),
                ctrl.get("control_type", ""),
                ctrl.get("title", ""),
                ctrl.get("description", ""),
                ctrl.get("control_owner", ""),
                ctrl.get("control_performer", ""),
                ctrl.get("frequency", ""),
                str(ctrl.get("page_number", "")),
                ctrl.get("section_heading", ""),
                ctrl.get("source_excerpt", ""),
                ctrl.get("expected_evidence", ""),
                match_label,
                protecht_id,
                "Open",
            ])

    _style_header_row(ws_controls, len(draft_headers))
    _wrap_rows(ws_controls)
    _auto_width(ws_controls)

    # -------------------------------------------------------------------------
    # Tab 2: Recommendations
    # -------------------------------------------------------------------------
    ws_gap = wb.create_sheet("Gap Analysis")

    gap_headers = [
        "Risk Event ID", "Localized Risk Event Name", "Localized Risk Event",
        "Coverage Tier", "Gap Description",
        "Protecht ID", "Control ID", "Control Name", "Control Description",
        "Control Owner", "Control Performer", "Control Type", "Frequency",
    ]
    ws_gap.append(gap_headers)

    tier_labels = {
        1: "Tier 1 — No Coverage",
        2: "Tier 2 — Partially Adequate Coverage",
        3: "Tier 3 — Adequate Coverage",
    }
    tier_fills = {
        1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        2: PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        3: PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    }

    # Build recommendation lookup by risk_event_id for easy access
    recs_by_reid = {}
    for rec in recommendations:
        recs_by_reid.setdefault(rec.get("risk_event_id", ""), []).append(rec)

    for finding in gap_findings:
        tier = finding.get("coverage_tier")
        tier_label = tier_labels.get(tier, "")
        risk_event_id = finding.get("risk_event_id", "")
        risk_name = finding.get("risk_name", "")
        risk_statement = finding.get("risk_statement", "")
        gap_description = finding.get("gap_description", "")
        controls_str = finding.get("controls_addressing", "None")

        associated_ids = [
            c.strip() for c in controls_str.split(",")
            if c.strip() and c.strip().lower() != "none"
        ]

        # Build one detail row per control
        ctrl_rows = []
        for ctrl_id in associated_ids:
            ctrl = controls_by_id.get(ctrl_id)
            if ctrl:
                ctrl_rows.append([
                    ctrl.get("repository_protecht_id", ""),
                    ctrl.get("control_id", ctrl_id),
                    ctrl.get("title", ""),
                    ctrl.get("description", ""),
                    ctrl.get("control_owner", ""),
                    ctrl.get("control_performer", ""),
                    ctrl.get("control_type", ""),
                    ctrl.get("frequency", ""),
                ])
            else:
                repo = repo_by_id.get(ctrl_id, {})
                ctrl_rows.append([
                    ctrl_id,
                    "",
                    repo.get("control_name", ""),
                    repo.get("control_description", ""),
                    "",
                    "",
                    repo.get("control_type", ""),
                    "",
                ])

        # Include draft recommendations for this risk
        for rec in recs_by_reid.get(risk_event_id, []):
            ctrl_rows.append([
                "",
                rec.get("recommendation_id", ""),
                rec.get("title", ""),
                rec.get("description", ""),
                rec.get("control_owner", "TBD"),
                rec.get("control_performer", "TBD"),
                rec.get("control_type", ""),
                rec.get("frequency", "TBD"),
            ])

        # If still no rows, write a single placeholder row
        if not ctrl_rows:
            ctrl_rows.append(["", "", "", "", "", "", "", ""])

        for ctrl_row in ctrl_rows:
            row_idx = ws_gap.max_row + 1
            ws_gap.append([
                risk_event_id,
                risk_name,
                risk_statement,
                tier_label,
                gap_description,
            ] + ctrl_row)
            if tier in tier_fills:
                ws_gap.cell(row=row_idx, column=4).fill = tier_fills[tier]
                ws_gap.cell(row=row_idx, column=4).font = Font(bold=True)

    _style_header_row(ws_gap, len(gap_headers))
    _wrap_rows(ws_gap)
    _auto_width(ws_gap)

    # -------------------------------------------------------------------------
    # Tab 3: Recommendations
    # -------------------------------------------------------------------------

    ws_recs = wb.create_sheet("Recommendations")

    rec_headers = [
        "Recommendation ID", "Risk Event ID", "Gap Tier", "Localized Risk Event Name", "Localized Risk Event",
        "Recommended Type", "Recommended Title", "Recommended Description",
        "Recommended Owner", "Recommended Performer", "Recommended Frequency", "Recommended Evidence",
        "Validation Status", "SME Notes",
    ]
    ws_recs.append(rec_headers)

    for rec in recommendations:
        ws_recs.append([
            rec.get("recommendation_id", ""),
            rec.get("risk_event_id", ""),
            f"Tier {rec.get('gap_tier', '')}",
            rec.get("risk_name", ""),
            rec.get("risk_statement", ""),
            rec.get("control_type", ""),
            rec.get("title", ""),
            rec.get("description", ""),
            rec.get("control_owner", "TBD"),
            rec.get("control_performer", "TBD"),
            rec.get("frequency", "TBD"),
            rec.get("expected_evidence", ""),
            "Draft — Requires SME Validation",
            "",
        ])

    _style_header_row(ws_recs, len(rec_headers))
    _wrap_rows(ws_recs)
    _auto_width(ws_recs)

    # -------------------------------------------------------------------------
    # Save
    # -------------------------------------------------------------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in document_name)
    filename = f"Controls - {safe_name} - {timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    return filepath
